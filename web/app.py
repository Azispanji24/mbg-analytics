# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

"""
=========================================================
SISTEM ANALISIS MENU MBG - FLASK BACKEND
=========================================================
Mode:
  1. PRESET  — muat hasil analisis notebook langsung (instant)
  2. DRIVE   — download gambar dari Google Drive + OCR + Mining (lama)
"""

import os, json, threading
from flask import Flask, jsonify, render_template, request
from flask_cors import CORS

app = Flask(__name__)
CORS(app)

# ─── Config ────────────────────────────────────────────────────────────────────
FOLDER_ID   = "1qXPs8jT-7lLpJqIN2m3bUVE-TKDUVCTv"
BASE_DIR    = os.path.dirname(__file__)
DATASET_DIR = os.path.join(BASE_DIR, "dataset")
CACHE_FILE  = os.path.join(BASE_DIR, "results_cache.json")

# ─── State ─────────────────────────────────────────────────────────────────────
processing_status = {
    "stage": "idle",
    "progress": 0,
    "message": "Pilih mode analisis",
    "total": 0,
    "current": 0,
    "error": None
}
results_cache = {}

# ─── Data Preset (hasil analisis notebook — 846 gambar) ────────────────────────
PRESET_DATA = {
    "meta": {
        "total_images": 846,
        "total_transactions": 830,
        "total_frequent_itemsets": 101,
        "total_rules": 56,
        "total_rules_filtered": 5,
        "unique_items": [
            "abon","apel","ayam","ikan","jamur","jeruk","kacang","kentang",
            "kurma","melon","mie","nasi","pisang","rendang","roti","sayur",
            "semangka","sup","susu","tahu","telur","tempe","ubi"
        ],
        "format_dist": {"png": 444, "jpg": 245, "jpeg": 157},
        "source": "preset"
    },
    "item_frequency": {
        "nasi":694,"telur":335,"ayam":319,"tahu":247,"sayur":245,
        "jeruk":237,"susu":229,"tempe":215,"pisang":205,"roti":176,
        "kacang":154,"ikan":128,"apel":123,"semangka":118,"melon":107,
        "kentang":98,"ubi":89,"jamur":76,"abon":71,"sup":71,
        "rendang":58,"kurma":52,"mie":48
    },
    "item_dist_per_transaction": {
        "1":37,"2":57,"3":154,"4":256,"5":197,"6":80,"7":30,"8":14,"9":3,"10":2
    },
    "frequent_itemsets": [
        {"itemset":"nasi","support":0.8361,"length":1},
        {"itemset":"telur","support":0.4036,"length":1},
        {"itemset":"ayam","support":0.3843,"length":1},
        {"itemset":"nasi, telur","support":0.3313,"length":2},
        {"itemset":"nasi, ayam","support":0.3289,"length":2},
        {"itemset":"tahu","support":0.2976,"length":1},
        {"itemset":"sayur","support":0.2952,"length":1},
        {"itemset":"sayur, nasi","support":0.2855,"length":2},
        {"itemset":"jeruk","support":0.2855,"length":1},
        {"itemset":"nasi, tahu","support":0.2783,"length":2},
        {"itemset":"susu","support":0.2759,"length":1},
        {"itemset":"tempe","support":0.2590,"length":1},
        {"itemset":"jeruk, nasi","support":0.2578,"length":2},
        {"itemset":"tempe, nasi","support":0.2482,"length":2},
        {"itemset":"pisang","support":0.2470,"length":1},
        {"itemset":"pisang, nasi","support":0.2145,"length":2},
        {"itemset":"roti","support":0.2120,"length":1},
        {"itemset":"nasi, susu","support":0.2000,"length":2},
        {"itemset":"kacang","support":0.1855,"length":1},
        {"itemset":"telur, susu","support":0.1747,"length":2},
        {"itemset":"ikan","support":0.1542,"length":1},
        {"itemset":"apel","support":0.1410,"length":1},
        {"itemset":"semangka","support":0.1422,"length":1},
        {"itemset":"nasi, ayam, telur","support":0.1494,"length":3},
        {"itemset":"nasi, tahu, sayur","support":0.1398,"length":3},
        {"itemset":"nasi, roti","support":0.1373,"length":2},
        {"itemset":"roti, telur","support":0.1277,"length":2},
        {"itemset":"roti, susu","support":0.1289,"length":2},
        {"itemset":"nasi, ikan","support":0.1229,"length":2},
        {"itemset":"nasi, jeruk, tempe","support":0.1157,"length":3},
        {"itemset":"tempe, sayur","support":0.1169,"length":2},
        {"itemset":"nasi, apel","support":0.1036,"length":2},
        {"itemset":"nasi, pisang, telur","support":0.1036,"length":3},
        {"itemset":"nasi, susu, telur","support":0.1289,"length":3},
        {"itemset":"nasi, tempe, sayur","support":0.1000,"length":3},
        {"itemset":"kentang","support":0.1181,"length":1},
        {"itemset":"nasi, kentang","support":0.0795,"length":2},
        {"itemset":"nasi, semangka","support":0.0988,"length":2},
        {"itemset":"tempe, telur","support":0.0904,"length":2},
        {"itemset":"nasi, roti, telur","support":0.0867,"length":3},
        {"itemset":"jeruk, nasi, tempe","support":0.0867,"length":3},
        {"itemset":"jeruk, sayur","support":0.0771,"length":2},
        {"itemset":"abon","support":0.0855,"length":1},
        {"itemset":"roti, ayam","support":0.0687,"length":2},
        {"itemset":"pisang, susu","support":0.0627,"length":2},
        {"itemset":"tempe, sayur, ayam","support":0.0602,"length":3},
        {"itemset":"pisang, tempe","support":0.0590,"length":2},
        {"itemset":"roti, nasi, susu","support":0.0506,"length":3},
        {"itemset":"abon, roti","support":0.0518,"length":2},
        {"itemset":"jeruk, susu","support":0.0530,"length":2},
        {"itemset":"ayam, susu","support":0.0506,"length":2},
        {"itemset":"jeruk, ayam, tempe","support":0.0542,"length":3},
        {"itemset":"tempe, telur, nasi","support":0.0904,"length":3},
        {"itemset":"roti, ayam, telur","support":0.0506,"length":3},
        {"itemset":"jeruk, nasi, tempe, ayam","support":0.0542,"length":4},
        {"itemset":"tempe, sayur, ayam, nasi","support":0.0602,"length":4},
    ],
    "rules": [
        {"antecedents":"abon","consequents":"roti","support":0.0518,"confidence":0.7288,"lift":3.4370,"leverage":0.0367},
        {"antecedents":"roti, telur","consequents":"susu","support":0.0795,"confidence":0.6226,"lift":2.2567,"leverage":0.0443},
        {"antecedents":"roti","consequents":"susu","support":0.1289,"confidence":0.6080,"lift":2.2035,"leverage":0.0704},
        {"antecedents":"nasi, apel","consequents":"susu","support":0.0795,"confidence":0.7669,"lift":2.7804,"leverage":0.0509},
        {"antecedents":"apel","consequents":"susu","support":0.0867,"confidence":0.6149,"lift":2.2293,"leverage":0.0478},
        {"antecedents":"melon","consequents":"pisang","support":0.0590,"confidence":0.5517,"lift":2.2335,"leverage":0.0326},
        {"antecedents":"nasi, ikan","consequents":"tahu","support":0.0795,"confidence":0.6471,"lift":2.1757,"leverage":0.0430},
        {"antecedents":"ikan","consequents":"tahu","support":0.0867,"confidence":0.5620,"lift":1.8895,"leverage":0.0408},
        {"antecedents":"roti, ayam","consequents":"telur","support":0.0506,"confidence":0.7368,"lift":1.8256,"leverage":0.0229},
        {"antecedents":"kurma","consequents":"pisang","support":0.0530,"confidence":0.6731,"lift":2.7248,"leverage":0.0335},
        {"antecedents":"ayam, susu","consequents":"telur","support":0.0506,"confidence":0.7000,"lift":1.7343,"leverage":0.0214},
        {"antecedents":"semangka","consequents":"pisang","support":0.0687,"confidence":0.4831,"lift":1.9561,"leverage":0.0336},
        {"antecedents":"pisang, susu","consequents":"telur","support":0.0627,"confidence":0.6933,"lift":1.7178,"leverage":0.0262},
        {"antecedents":"jeruk, susu","consequents":"telur","support":0.0530,"confidence":0.6875,"lift":1.7034,"leverage":0.0219},
        {"antecedents":"nasi, tahu","consequents":"sayur","support":0.1398,"confidence":0.5028,"lift":1.7036,"leverage":0.0577},
        {"antecedents":"nasi, susu","consequents":"telur","support":0.1289,"confidence":0.6446,"lift":1.5970,"leverage":0.0482},
        {"antecedents":"jeruk, nasi, tempe","consequents":"ayam","support":0.0542,"confidence":0.6081,"lift":1.5822,"leverage":0.0200},
        {"antecedents":"susu","consequents":"telur","support":0.1747,"confidence":0.6332,"lift":1.5688,"leverage":0.0633},
        {"antecedents":"roti, nasi","consequents":"telur","support":0.0867,"confidence":0.6316,"lift":1.5648,"leverage":0.0314},
        {"antecedents":"roti, nasi, susu","consequents":"telur","support":0.0506,"confidence":0.6269,"lift":1.5531,"leverage":0.0181},
        {"antecedents":"roti, susu","consequents":"telur","support":0.0795,"confidence":0.6168,"lift":1.5282,"leverage":0.0275},
        {"antecedents":"roti","consequents":"telur","support":0.1277,"confidence":0.6023,"lift":1.4922,"leverage":0.0421},
        {"antecedents":"tempe, nasi","consequents":"sayur","support":0.1000,"confidence":0.4030,"lift":1.3657,"leverage":0.0268},
        {"antecedents":"jeruk, ayam, tempe","consequents":"nasi","support":0.0542,"confidence":1.0000,"lift":1.1960,"leverage":0.0089},
        {"antecedents":"tempe, sayur, ayam","consequents":"nasi","support":0.0602,"confidence":1.0000,"lift":1.1960,"leverage":0.0099},
        {"antecedents":"tempe, telur","consequents":"nasi","support":0.0904,"confidence":1.0000,"lift":1.1960,"leverage":0.0148},
        {"antecedents":"tempe, sayur","consequents":"nasi","support":0.1169,"confidence":0.9898,"lift":1.1838,"leverage":0.0182},
        {"antecedents":"jeruk, sayur","consequents":"nasi","support":0.0771,"confidence":0.9846,"lift":1.1776,"leverage":0.0117},
        {"antecedents":"pisang, tempe","consequents":"nasi","support":0.0590,"confidence":0.9800,"lift":1.1720,"leverage":0.0087},
        {"antecedents":"ayam, telur","consequents":"nasi","support":0.1494,"confidence":0.9615,"lift":1.1497,"leverage":0.0196},
        {"antecedents":"nasi, ayam","consequents":"telur","support":0.1494,"confidence":0.4545,"lift":1.1263,"leverage":0.0167},
        {"antecedents":"nasi, tahu","consequents":"ayam","support":0.0795,"confidence":0.2857,"lift":0.7432,"leverage":-0.0274},
        {"antecedents":"nasi","consequents":"ayam","support":0.3289,"confidence":0.3934,"lift":1.0235,"leverage":0.0076},
        {"antecedents":"nasi","consequents":"tahu","support":0.2783,"confidence":0.3328,"lift":1.1183,"leverage":0.0292},
        {"antecedents":"nasi","consequents":"sayur","support":0.2855,"confidence":0.3414,"lift":1.1566,"leverage":0.0386},
        {"antecedents":"nasi","consequents":"telur","support":0.3313,"confidence":0.3961,"lift":0.9813,"leverage":-0.0065},
        {"antecedents":"nasi","consequents":"susu","support":0.2000,"confidence":0.2392,"lift":0.8668,"leverage":-0.0308},
        {"antecedents":"nasi","consequents":"jeruk","support":0.2578,"confidence":0.3083,"lift":1.0796,"leverage":0.0193},
        {"antecedents":"nasi","consequents":"tempe","support":0.2482,"confidence":0.2968,"lift":1.1457,"leverage":0.0314},
        {"antecedents":"telur","consequents":"nasi","support":0.3313,"confidence":0.8209,"lift":0.9813,"leverage":-0.0065},
        {"antecedents":"ayam","consequents":"nasi","support":0.3289,"confidence":0.8558,"lift":1.0235,"leverage":0.0076},
        {"antecedents":"tahu","consequents":"nasi","support":0.2783,"confidence":0.9355,"lift":1.1183,"leverage":0.0292},
        {"antecedents":"sayur","consequents":"nasi","support":0.2855,"confidence":0.9680,"lift":1.1566,"leverage":0.0386},
        {"antecedents":"susu","consequents":"nasi","support":0.2000,"confidence":0.7249,"lift":0.8668,"leverage":-0.0308},
        {"antecedents":"jeruk","consequents":"nasi","support":0.2578,"confidence":0.9025,"lift":1.0796,"leverage":0.0193},
        {"antecedents":"tempe","consequents":"nasi","support":0.2482,"confidence":0.9580,"lift":1.1457,"leverage":0.0314},
        {"antecedents":"pisang","consequents":"nasi","support":0.2145,"confidence":0.8684,"lift":1.0385,"leverage":0.0082},
        {"antecedents":"roti","consequents":"nasi","support":0.1373,"confidence":0.6477,"lift":0.7749,"leverage":-0.0400},
        {"antecedents":"kacang","consequents":"nasi","support":0.1518,"confidence":0.8189,"lift":0.9794,"leverage":-0.0032},
        {"antecedents":"abon","consequents":"nasi","support":0.0687,"confidence":0.8028,"lift":0.9602,"leverage":-0.0028},
        {"antecedents":"ikan","consequents":"nasi","support":0.1229,"confidence":0.7973,"lift":0.9534,"leverage":-0.0059},
        {"antecedents":"apel","consequents":"nasi","support":0.1036,"confidence":0.7345,"lift":0.8786,"leverage":-0.0143},
        {"antecedents":"semangka","consequents":"nasi","support":0.0988,"confidence":0.6949,"lift":0.8314,"leverage":-0.0201},
        {"antecedents":"kentang","consequents":"nasi","support":0.0795,"confidence":0.6726,"lift":0.8046,"leverage":-0.0193},
        {"antecedents":"roti, ayam","consequents":"nasi","support":0.0506,"confidence":0.7368,"lift":0.8813,"leverage":-0.0068},
        {"antecedents":"pisang, susu","consequents":"nasi","support":0.0627,"confidence":0.6933,"lift":0.8297,"leverage":-0.0128},
    ],
    "rules_filtered": [
        {"antecedents":"abon","consequents":"roti","support":0.0518,"confidence":0.7288,"lift":3.4370},
        {"antecedents":"roti, ayam","consequents":"telur","support":0.0506,"confidence":0.7368,"lift":1.8256},
        {"antecedents":"ayam, susu","consequents":"telur","support":0.0506,"confidence":0.7000,"lift":1.7343},
        {"antecedents":"pisang, susu","consequents":"telur","support":0.0627,"confidence":0.6933,"lift":1.7178},
        {"antecedents":"jeruk, susu","consequents":"telur","support":0.0530,"confidence":0.6875,"lift":1.7034},
    ]
}

# ─── Normalisasi item ───────────────────────────────────────────────────────────
VALID_FOOD_ITEMS = {
    "nasi","ayam","telur","tahu","tempe","sayur","ikan","susu","roti","mie",
    "kentang","wortel","bayam","kacang","pisang","jeruk","apel","semangka",
    "melon","jamur","sup","rendang","abon","ubi","kurma",
}
SYNONYMS = {
    "rice":"nasi","chicken":"ayam","egg":"telur","tofu":"tahu","vegetable":"sayur",
    "sayuran":"sayur","fish":"ikan","milk":"susu","bread":"roti","noodle":"mie",
    "banana":"pisang","orange":"jeruk","apple":"apel","watermelon":"semangka",
    "potato":"kentang","spinach":"bayam","carrot":"wortel","peanut":"kacang",
    "mushroom":"jamur","soup":"sup","beef":"rendang","yam":"ubi","dates":"kurma",
}

def normalize_item(word):
    w = word.lower().strip()
    if w in VALID_FOOD_ITEMS: return w
    if w in SYNONYMS: return SYNONYMS[w]
    for item in VALID_FOOD_ITEMS:
        if item in w or w in item: return item
    return None

def extract_items(ocr_words):
    found = set()
    for word in ocr_words:
        norm = normalize_item(word)
        if norm: found.add(norm)
    return sorted(found)

# ─── Load cache dari disk ───────────────────────────────────────────────────────
def load_cache():
    global results_cache, processing_status
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, "r", encoding="utf-8") as f:
            results_cache = json.load(f)
        processing_status.update({
            "stage": "done", "progress": 100,
            "message": "Data berhasil dimuat dari cache"
        })
        print("[OK] Cache dimuat:", CACHE_FILE)

load_cache()

# ─── Pipeline: Mode PRESET (instant) ───────────────────────────────────────────
def run_preset():
    """Muat langsung data hasil analisis notebook — tidak perlu download."""
    global processing_status, results_cache
    try:
        steps = [
            (10,  "Memuat metadata dataset..."),
            (30,  "Memuat data transaksi (830 gambar valid)..."),
            (55,  "Memuat frequent itemsets (101 itemset)..."),
            (75,  "Memuat association rules (56 rules)..."),
            (90,  "Memfilter rules signifikan..."),
            (96,  "Menyimpan ke cache lokal..."),
        ]
        import time
        for pct, msg in steps:
            processing_status.update({"progress": pct, "message": msg})
            time.sleep(0.4)

        results_cache = PRESET_DATA.copy()
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(results_cache, f, ensure_ascii=False, indent=2)

        processing_status.update({
            "stage": "done", "progress": 100,
            "message": "Data preset berhasil dimuat! 56 rules dari 846 gambar."
        })
        print("[OK] Preset data loaded.")
    except Exception as e:
        import traceback
        processing_status.update({
            "stage": "error", "error": str(e),
            "message": f"Error preset: {e}"
        })
        print("[ERROR]", traceback.format_exc())

# ─── Pipeline: Mode DRIVE (download incremental + OCR cache + mining) ──────────
def run_drive_pipeline():
    """Sinkronisasi dengan Drive (hanya gambar baru), OCR dengan cache, lalu FP-Growth."""
    global processing_status, results_cache
    try:
        import time
        img_exts = (".jpg", ".jpeg", ".png")
        os.makedirs(DATASET_DIR, exist_ok=True)

        # ── Tahap 1: Sinkronisasi dengan Google Drive ────────────────────────
        processing_status.update({
            "stage": "downloading", "progress": 3,
            "message": "Menghubungi Google Drive...", "error": None
        })

        # Hitung file lokal saat ini
        existing_local = set(
            f for f in os.listdir(DATASET_DIR) if f.lower().endswith(img_exts)
        )
        n_existing = len(existing_local)
        processing_status.update({
            "progress": 8,
            "message": f"{n_existing} gambar lokal ditemukan. Menyinkronkan..."
        })

        # gdown download folder — otomatis skip file yang sudah ada
        try:
            import gdown
            gdown.download_folder(
                f"https://drive.google.com/drive/folders/{FOLDER_ID}",
                output=DATASET_DIR,
                quiet=True,
                use_cookies=False,
            )
        except Exception as dl_err:
            print(f"[WARN] gdown: {dl_err} — lanjut dengan file lokal")

        # Scan semua gambar (rekursif jika ada subfolder dari gdown)
        image_files = []
        for root, _, files in os.walk(DATASET_DIR):
            for fname in sorted(files):
                if fname.lower().endswith(img_exts):
                    rel = os.path.relpath(os.path.join(root, fname), DATASET_DIR)
                    image_files.append(rel)

        total_images = len(image_files)
        new_count = max(0, total_images - n_existing)
        processing_status.update({
            "progress": 18, "total": total_images,
            "message": f"Sinkronisasi OK: {total_images} gambar total, {new_count} baru"
        })
        time.sleep(0.4)

        # ── Tahap 2: OCR dengan cache per-file ──────────────────────────────
        OCR_CACHE_FILE = os.path.join(BASE_DIR, "ocr_cache.json")
        ocr_cache = {}
        if os.path.exists(OCR_CACHE_FILE):
            with open(OCR_CACHE_FILE, "r", encoding="utf-8") as fc:
                ocr_cache = json.load(fc)

        n_cached = sum(1 for f in image_files if f in ocr_cache)
        n_todo = total_images - n_cached

        processing_status.update({
            "stage": "ocr", "progress": 20,
            "message": f"OCR: {n_cached} dari cache, {n_todo} harus di-proses baru..."
        })

        if n_todo > 0:
            import easyocr
            reader = easyocr.Reader(["id", "en"], gpu=False)
        else:
            reader = None

        ocr_results = {}
        newly_ocr = 0
        for i, fname in enumerate(image_files):
            img_path = os.path.join(DATASET_DIR, fname)
            if fname in ocr_cache:
                ocr_results[fname] = ocr_cache[fname]
                label = "[cache]"
            else:
                try:
                    ocr_out = reader.readtext(img_path, detail=0)
                    ocr_results[fname] = [str(t) for t in ocr_out]
                except Exception:
                    ocr_results[fname] = []
                ocr_cache[fname] = ocr_results[fname]
                newly_ocr += 1
                label = "[baru]"
                # Simpan cache checkpoint tiap 25 file baru
                if newly_ocr % 25 == 0:
                    with open(OCR_CACHE_FILE, "w", encoding="utf-8") as fc:
                        json.dump(ocr_cache, fc, ensure_ascii=False)

            pct = 20 + int((i + 1) / total_images * 52)
            bname = os.path.basename(fname)
            processing_status.update({
                "progress": pct, "current": i + 1,
                "message": f"OCR {i+1}/{total_images} {label} — {bname[:40]}"
            })

        # Simpan OCR cache final
        with open(OCR_CACHE_FILE, "w", encoding="utf-8") as fc:
            json.dump(ocr_cache, fc, ensure_ascii=False)
        print(f"[INFO] OCR selesai: {newly_ocr} baru, {total_images - newly_ocr} cache")

        # ── Tahap 3: Transaksi ───────────────────────────────────────────────
        processing_status.update({
            "stage": "mining", "progress": 74,
            "message": "Membangun data transaksi..."
        })
        transaction_list = [
            extract_items(w)
            for w in ocr_results.values()
            if extract_items(w)
        ]

        # ── Tahap 4: FP-Growth + Association Rules ───────────────────────────
        processing_status.update({"progress": 80, "message": "One-Hot Encoding..."})
        from mlxtend.preprocessing import TransactionEncoder
        from mlxtend.frequent_patterns import fpgrowth, association_rules
        import pandas as pd
        from collections import Counter

        te = TransactionEncoder()
        te_arr = te.fit(transaction_list).transform(transaction_list)
        df = pd.DataFrame(te_arr, columns=te.columns_)

        processing_status.update({"progress": 86, "message": "Menjalankan FP-Growth..."})
        freq_items = fpgrowth(df, min_support=0.05, use_colnames=True)
        freq_items["length"] = freq_items["itemsets"].apply(len)
        freq_items = freq_items.sort_values("support", ascending=False).reset_index(drop=True)

        processing_status.update({"progress": 92, "message": "Generate Association Rules..."})
        rules = association_rules(freq_items, metric="confidence", min_threshold=0.6)
        rules = rules.sort_values("lift", ascending=False).reset_index(drop=True)
        rules_filtered = rules[(rules["lift"] > 1.5) & (rules["confidence"] > 0.65)].copy()

        # ── Serialize ────────────────────────────────────────────────────────
        def s2str(s): return ", ".join(sorted(s))
        item_freq = Counter(item for t in transaction_list for item in t)
        item_dist = Counter(len(t) for t in transaction_list)
        fmt_dist = {"png": 0, "jpg": 0, "jpeg": 0}
        for f in image_files:
            ext = os.path.basename(f).lower().rsplit(".", 1)[-1]
            if ext in fmt_dist:
                fmt_dist[ext] += 1

        from datetime import datetime
        results_cache = {
            "meta": {
                "total_images": total_images,
                "total_transactions": len(transaction_list),
                "total_frequent_itemsets": len(freq_items),
                "total_rules": len(rules),
                "total_rules_filtered": len(rules_filtered),
                "unique_items": sorted(df.columns.tolist()),
                "format_dist": fmt_dist,
                "source": "drive",
                "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "new_images_this_run": new_count,
            },
            "item_frequency": dict(item_freq.most_common(30)),
            "item_dist_per_transaction": {str(k): v for k, v in sorted(item_dist.items())},
            "frequent_itemsets": [
                {"itemset": s2str(r["itemsets"]), "support": round(float(r["support"]), 4),
                 "length": int(r["length"])}
                for _, r in freq_items.iterrows()
            ],
            "rules": [
                {"antecedents": s2str(r["antecedents"]), "consequents": s2str(r["consequents"]),
                 "support": round(float(r["support"]), 4), "confidence": round(float(r["confidence"]), 4),
                 "lift": round(float(r["lift"]), 4), "leverage": round(float(r["leverage"]), 4)}
                for _, r in rules.iterrows()
            ],
            "rules_filtered": [
                {"antecedents": s2str(r["antecedents"]), "consequents": s2str(r["consequents"]),
                 "support": round(float(r["support"]), 4), "confidence": round(float(r["confidence"]), 4),
                 "lift": round(float(r["lift"]), 4)}
                for _, r in rules_filtered.iterrows()
            ],
        }

        processing_status.update({"progress": 98, "message": "Menyimpan hasil..."})
        with open(CACHE_FILE, "w", encoding="utf-8") as fout:
            json.dump(results_cache, fout, ensure_ascii=False, indent=2)
        print("[OK] Drive pipeline selesai.")

        processing_status.update({
            "stage": "done", "progress": 100,
            "message": f"Analisis selesai! {len(rules)} rules dari {total_images} gambar."
        })
    except Exception as e:
        import traceback
        processing_status.update({
            "stage": "error", "error": str(e), "message": f"Error: {e}"
        })
        print("[ERROR]", traceback.format_exc())

# ─── Routes ────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/status")
def api_status():
    return jsonify(processing_status)

@app.route("/api/start-preset", methods=["POST"])
def api_start_preset():
    """Mode cepat: muat data hasil analisis notebook (instant ~2 detik)."""
    if processing_status["stage"] in ("downloading","ocr","mining"):
        return jsonify({"ok": False, "msg": "Pemrosesan sedang berjalan"}), 409
    processing_status.update({"stage":"mining","progress":5,"message":"Memuat data preset...","error":None})
    threading.Thread(target=run_preset, daemon=True).start()
    return jsonify({"ok": True, "msg": "Mode preset dimulai"})

@app.route("/api/start-drive", methods=["POST"])
def api_start_drive():
    """Mode lengkap: download dari Google Drive + OCR + Mining (30+ menit)."""
    if processing_status["stage"] in ("downloading","ocr","mining"):
        return jsonify({"ok": False, "msg": "Pemrosesan sedang berjalan"}), 409
    processing_status.update({"stage":"downloading","progress":5,"message":"Memulai download dari Drive...","error":None})
    threading.Thread(target=run_drive_pipeline, daemon=True).start()
    return jsonify({"ok": True, "msg": "Mode Drive dimulai"})

@app.route("/api/reset", methods=["POST"])
def api_reset():
    global results_cache
    results_cache = {}
    processing_status.update({
        "stage":"idle","progress":0,"current":0,"total":0,
        "message":"Pilih mode analisis","error":None
    })
    if os.path.exists(CACHE_FILE):
        os.remove(CACHE_FILE)
    return jsonify({"ok": True})

@app.route("/api/results")
def api_results():
    if not results_cache:
        return jsonify({"ok": False, "msg": "Belum ada hasil"}), 404
    return jsonify({"ok": True, "data": results_cache})

@app.route("/api/rules")
def api_rules():
    if not results_cache:
        return jsonify({"ok": False}), 404
    limit    = int(request.args.get("limit", 50))
    min_lift = float(request.args.get("min_lift", 0))
    min_conf = float(request.args.get("min_conf", 0))
    rules = [
        r for r in results_cache.get("rules", [])
        if r["lift"] >= min_lift and r["confidence"] >= min_conf
    ][:limit]
    return jsonify({"ok": True, "rules": rules, "total": len(rules)})

@app.route("/api/download")
def api_download():
    """Hasilkan file Excel (.xlsx) multi-sheet berisi semua hasil analisis."""
    if not results_cache:
        return jsonify({"ok": False, "msg": "Belum ada hasil analisis"}), 404

    import io
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    meta = results_cache.get("meta", {})
    ts   = meta.get("last_updated", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    wb = Workbook()
    wb.remove(wb.active)  # hapus sheet default kosong

    # ── Palette warna ────────────────────────────────────────────────────
    CLR_HEADER_DARK  = "1E1B4B"   # ungu gelap   → header utama
    CLR_HEADER_MID   = "312E81"   # ungu sedang  → sub-header
    CLR_ACCENT_BLUE  = "0EA5E9"   # biru         → Rules Terbaik header
    CLR_ACCENT_GREEN = "166534"   # hijau gelap  → Itemsets header
    CLR_ROW_ALT      = "F1F5FF"   # biru pucat   → baris alternating
    CLR_ROW_TOP      = "FEF9C3"   # kuning pucat → top rules highlight
    CLR_WHITE        = "FFFFFF"
    CLR_TEXT_LIGHT   = "F8FAFC"

    thin_side  = Side(style="thin",   color="D1D5DB")
    thick_side = Side(style="medium", color="6C63FF")

    def header_font(color=CLR_TEXT_LIGHT, size=11, bold=True):
        return Font(name="Calibri", bold=bold, color=color, size=size)

    def body_font(bold=False, color="111827", size=10):
        return Font(name="Calibri", bold=bold, color=color, size=size)

    def make_header_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def cell_border():
        return Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

    def apply_header_row(ws, row_idx, headers, fill_color=CLR_HEADER_DARK):
        fill = make_header_fill(fill_color)
        fnt  = header_font()
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=h)
            c.fill      = fill
            c.font      = fnt
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = cell_border()

    def apply_data_row(ws, row_idx, values, alt=False, highlight_color=None):
        fc = highlight_color or (CLR_ROW_ALT if alt else CLR_WHITE)
        fill = PatternFill("solid", fgColor=fc)
        fnt  = body_font()
        for col_idx, v in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=v)
            c.fill      = fill
            c.font      = fnt
            c.alignment = Alignment(vertical="center", wrap_text=False)
            c.border    = cell_border()

    def auto_col_width(ws, min_w=10, max_w=50):
        for col_cells in ws.columns:
            length = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in col_cells
            )
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max_w, max(min_w, length + 4))

    def add_title_cell(ws, title, subtitle=""):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        c = ws.cell(row=1, column=1, value=title)
        c.font      = Font(name="Calibri", bold=True, size=14, color=CLR_TEXT_LIGHT)
        c.fill      = make_header_fill(CLR_HEADER_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26
        if subtitle:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
            c2 = ws.cell(row=2, column=1, value=subtitle)
            c2.font      = Font(name="Calibri", size=10, color="A0AEC0")
            c2.fill      = make_header_fill("1A1D2E")
            c2.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[2].height = 18
            return 3   # data mulai baris 3
        return 2       # data mulai baris 2

    # ════════════════════════════════════════════════════════════════════
    # SHEET 1 — Ringkasan (Summary)
    # ════════════════════════════════════════════════════════════════════
    ws_sum = wb.create_sheet("📊 Ringkasan")

    # Judul besar
    ws_sum.merge_cells("A1:D1")
    c = ws_sum["A1"]
    c.value     = "MBG ANALYTICS — RINGKASAN HASIL ANALISIS"
    c.font      = Font(name="Calibri", bold=True, size=14, color=CLR_TEXT_LIGHT)
    c.fill      = make_header_fill(CLR_HEADER_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 28

    ws_sum.merge_cells("A2:D2")
    c2 = ws_sum["A2"]
    c2.value     = "Kelompok 6 — Tugas Machine Learning"
    c2.font      = Font(name="Calibri", size=10, color="94A3B8")
    c2.fill      = make_header_fill("1A1D2E")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[2].height = 18

    # Label-Value
    total_rules_all = len(results_cache.get("rules", []))
    top_rule        = results_cache.get("rules", [{}])[0]
    rows_info = [
        ("", ""),
        ("INFORMASI UMUM", ""),
        ("Tanggal Analisis",   ts),
        ("Sumber Data",        meta.get("source", "-").upper()),
        ("Link Google Drive",  "https://drive.google.com/drive/folders/1qXPs8jT-7lLpJqIN2m3bUVE-TKDUVCTv"),
        ("", ""),
        ("DATASET", ""),
        ("Total Gambar",       meta.get("total_images", 0)),
        ("Transaksi Valid",    meta.get("total_transactions", 0)),
        ("Item Unik",          len(meta.get("unique_items", []))),
        ("Format PNG",         meta.get("format_dist", {}).get("png", 0)),
        ("Format JPG",         meta.get("format_dist", {}).get("jpg", 0)),
        ("Format JPEG",        meta.get("format_dist", {}).get("jpeg", 0)),
        ("", ""),
        ("HASIL FP-GROWTH (min_support = 5%)", ""),
        ("Frequent Itemsets",  meta.get("total_frequent_itemsets", 0)),
        ("", ""),
        ("ASSOCIATION RULES (min_confidence = 60%)", ""),
        ("Total Rules",        total_rules_all),
        ("Rules Signifikan",   meta.get("total_rules_filtered", 0)),
        ("Kriteria Signifikan","lift > 1.5 AND confidence > 65%"),
        ("", ""),
        ("TOP RULE TERBAIK", ""),
        ("Antecedents → Consequents", f"{top_rule.get('antecedents','-')} → {top_rule.get('consequents','-')}"),
        ("Support",   f"{top_rule.get('support',0)*100:.1f}%"),
        ("Confidence",f"{top_rule.get('confidence',0)*100:.1f}%"),
        ("Lift",      f"{top_rule.get('lift',0):.3f}"),
    ]

    for r_idx, (label, val) in enumerate(rows_info, 3):
        ws_sum.row_dimensions[r_idx].height = 18
        if val == "" and label == "":
            continue
        if val == "" and label != "":
            # Section header
            ws_sum.merge_cells(f"A{r_idx}:D{r_idx}")
            c = ws_sum.cell(row=r_idx, column=1, value=label)
            c.font      = Font(name="Calibri", bold=True, size=11, color=CLR_TEXT_LIGHT)
            c.fill      = make_header_fill(CLR_HEADER_MID)
            c.alignment = Alignment(vertical="center")
            c.border    = cell_border()
        else:
            cl = ws_sum.cell(row=r_idx, column=1, value=label)
            cv = ws_sum.cell(row=r_idx, column=2, value=val)
            for c in (cl, cv):
                c.font      = body_font(bold=(c.column == 1))
                c.alignment = Alignment(vertical="center")
                c.border    = cell_border()
                ws_sum.merge_cells(f"B{r_idx}:D{r_idx}")

    ws_sum.column_dimensions["A"].width = 36
    ws_sum.column_dimensions["B"].width = 55
    ws_sum.freeze_panes = "A3"

    # ════════════════════════════════════════════════════════════════════
    # SHEET 2 — Association Rules (semua)
    # ════════════════════════════════════════════════════════════════════
    ws_rules = wb.create_sheet("🔗 Association Rules")
    data_start = add_title_cell(
        ws_rules,
        "Association Rules — Semua (min_confidence = 60%, diurutkan lift ↓)",
        f"Total: {total_rules_all} rules  |  Sumber: {meta.get('source','-').upper()}  |  {ts}"
    )

    headers_rules = ["#", "Antecedents (Jika ada...)", "Consequents (...maka ada)",
                     "Support", "Confidence", "Lift", "Leverage", "Kekuatan"]
    apply_header_row(ws_rules, data_start, headers_rules)
    ws_rules.row_dimensions[data_start].height = 22

    for i, r in enumerate(results_cache.get("rules", []), 1):
        row_idx = data_start + i
        lift    = r.get("lift", 0)
        if lift >= 2.0:   strength = "🔥 Sangat Kuat"
        elif lift >= 1.5: strength = "⚡ Kuat"
        else:             strength = "📌 Lemah"
        hc = None if lift < 1.5 else (CLR_ROW_TOP if lift >= 2.0 else "FFF7ED")
        apply_data_row(ws_rules, row_idx, [
            i,
            r.get("antecedents", ""),
            r.get("consequents", ""),
            f"{r.get('support',0)*100:.2f}%",
            f"{r.get('confidence',0)*100:.2f}%",
            round(r.get("lift", 0), 4),
            round(r.get("leverage", 0), 4),
            strength,
        ], alt=(i % 2 == 0), highlight_color=hc)
        ws_rules.row_dimensions[row_idx].height = 18

    auto_col_width(ws_rules, min_w=8)
    ws_rules.freeze_panes = f"A{data_start + 1}"

    # ════════════════════════════════════════════════════════════════════
    # SHEET 3 — Rules Terbaik (filtered)
    # ════════════════════════════════════════════════════════════════════
    ws_top = wb.create_sheet("⭐ Rules Terbaik")
    top_rules = results_cache.get("rules_filtered", [])
    ds_top = add_title_cell(
        ws_top,
        "Rules Signifikan — Lift > 1.5 & Confidence > 65%",
        f"Total: {len(top_rules)} rules terpilih"
    )

    headers_top = ["#", "Antecedents (Jika ada...)", "Consequents (...maka ada)",
                   "Support", "Confidence", "Lift", "Interpretasi"]
    apply_header_row(ws_top, ds_top, headers_top, fill_color=CLR_ACCENT_BLUE)
    ws_top.row_dimensions[ds_top].height = 22

    interp_map = {
        "abon → roti":           "Jika ada abon, kemungkinan 72.9% ada roti (3.4x lebih sering)",
        "roti, ayam → telur":    "Paket roti+ayam hampir selalu disertai telur (73.7%)",
        "ayam, susu → telur":    "Kombinasi ayam+susu erat dengan telur (70.0%)",
        "pisang, susu → telur":  "Pisang+susu sering muncul bersama telur (69.3%)",
        "jeruk, susu → telur":   "Jeruk+susu berkorelasi kuat dengan telur (68.8%)",
    }

    for i, r in enumerate(top_rules, 1):
        row_idx = ds_top + i
        ant, con = r.get("antecedents", ""), r.get("consequents", "")
        key = f"{ant} → {con}"
        interp = interp_map.get(key, f"Lift {r.get('lift',0):.2f}x di atas baseline")
        apply_data_row(ws_top, row_idx, [
            i, ant, con,
            f"{r.get('support',0)*100:.2f}%",
            f"{r.get('confidence',0)*100:.2f}%",
            round(r.get("lift", 0), 4),
            interp,
        ], alt=(i % 2 == 0), highlight_color=("FFFBEB" if i % 2 == 0 else None))
        ws_top.row_dimensions[row_idx].height = 20

    auto_col_width(ws_top, min_w=8)
    ws_top.freeze_panes = f"A{ds_top + 1}"

    # ════════════════════════════════════════════════════════════════════
    # SHEET 4 — Frequent Itemsets
    # ════════════════════════════════════════════════════════════════════
    ws_fi = wb.create_sheet("📦 Frequent Itemsets")
    fi_list = results_cache.get("frequent_itemsets", [])
    ds_fi = add_title_cell(
        ws_fi,
        "Frequent Itemsets (min_support = 5%)",
        f"Total: {len(fi_list)} itemsets"
    )

    headers_fi = ["#", "Itemset", "Support (%)", "Support (raw)", "Jumlah Item"]
    apply_header_row(ws_fi, ds_fi, headers_fi, fill_color=CLR_ACCENT_GREEN)
    ws_fi.row_dimensions[ds_fi].height = 22

    for i, row in enumerate(fi_list, 1):
        row_idx = ds_fi + i
        sup = row.get("support", 0)
        apply_data_row(ws_fi, row_idx, [
            i,
            row.get("itemset", ""),
            f"{sup*100:.2f}%",
            round(sup, 4),
            row.get("length", 1),
        ], alt=(i % 2 == 0))
        ws_fi.row_dimensions[row_idx].height = 18

    auto_col_width(ws_fi, min_w=8)
    ws_fi.freeze_panes = f"A{ds_fi + 1}"

    # ════════════════════════════════════════════════════════════════════
    # SHEET 5 — Frekuensi Item
    # ════════════════════════════════════════════════════════════════════
    ws_freq = wb.create_sheet("🥗 Frekuensi Item")
    freq_dict = results_cache.get("item_frequency", {})
    freq_sorted = sorted(freq_dict.items(), key=lambda x: -x[1])
    total_tx = meta.get("total_transactions", 1) or 1

    ds_freq = add_title_cell(
        ws_freq,
        "Frekuensi Kemunculan Item Makanan",
        f"Dari {total_tx} transaksi valid"
    )

    headers_freq = ["#", "Item", "Frekuensi", "Persentase (%)", "Peringkat"]
    apply_header_row(ws_freq, ds_freq, headers_freq, fill_color=CLR_HEADER_DARK)
    ws_freq.row_dimensions[ds_freq].height = 22

    for i, (item, freq_val) in enumerate(freq_sorted, 1):
        row_idx = ds_freq + i
        pct = freq_val / total_tx * 100
        apply_data_row(ws_freq, row_idx, [
            i,
            item.capitalize(),
            freq_val,
            f"{pct:.1f}%",
            f"#{i}",
        ], alt=(i % 2 == 0))
        ws_freq.row_dimensions[row_idx].height = 18

    auto_col_width(ws_freq, min_w=8)
    ws_freq.freeze_panes = f"A{ds_freq + 1}"

    # ════════════════════════════════════════════════════════════════════
    # SHEET 6 — Distribusi Item per Transaksi
    # ════════════════════════════════════════════════════════════════════
    ws_dist = wb.create_sheet("📈 Distribusi Transaksi")
    dist_dict = results_cache.get("item_dist_per_transaction", {})
    dist_sorted = sorted(dist_dict.items(), key=lambda x: int(x[0]))
    total_tx_dist = sum(dist_dict.values())

    ds_dist = add_title_cell(
        ws_dist,
        "Distribusi Jumlah Item per Transaksi",
        f"Total {total_tx_dist} transaksi valid"
    )

    headers_dist = ["Jumlah Item", "Jumlah Transaksi", "Persentase (%)"]
    apply_header_row(ws_dist, ds_dist, headers_dist, fill_color=CLR_HEADER_DARK)
    ws_dist.row_dimensions[ds_dist].height = 22

    for i, (k, v) in enumerate(dist_sorted, 1):
        row_idx = ds_dist + i
        pct = v / total_tx_dist * 100 if total_tx_dist else 0
        apply_data_row(ws_dist, row_idx, [
            int(k), v, f"{pct:.1f}%"
        ], alt=(i % 2 == 0))
        ws_dist.row_dimensions[row_idx].height = 18

    auto_col_width(ws_dist, min_w=14)
    ws_dist.freeze_panes = f"A{ds_dist + 1}"

    # ── Simpan ke buffer ─────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_ts = ts.replace(":", "-").replace(" ", "_")
    fname   = f"MBG_Analytics_Hasil_{safe_ts}.xlsx"

    from flask import send_file
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname
    )

if __name__ == "__main__":
    app.run(debug=True, port=5000)

